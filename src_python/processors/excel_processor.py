"""
Excel Data Processor Module
Handles reading and extracting data from Excel files.
"""

import asyncio
from pathlib import Path
from typing import Dict, List, Any, Optional
import hashlib

import pandas as pd
from openpyxl import load_workbook

from ..utils.logger import get_logger
from ..utils.config import PipelineConfig


logger = get_logger("excel-processor")


class ExcelProcessor:
    """
    Processor for Excel files.
    """
    
    def __init__(self, config: PipelineConfig):
        """
        Initialize Excel processor.
        
        Args:
            config: Pipeline configuration
        """
        self.config = config
        self.chunk_size = config.chunk_size
        self.chunk_overlap = config.chunk_overlap
    
    async def read_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Read Excel file and extract data.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of sheet data
            
        Raises:
            FileNotFoundError: If file doesn't exist
            Exception: If reading fails
        """
        try:
            logger.info(f"Reading Excel file: {file_path}")
            
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")
            
            # Load workbook
            wb = load_workbook(file_path, read_only=True)
            result = []
            
            # Iterate through all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                
                # Read all rows
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                
                # Clean data
                cleaned_data = self._clean_data(data)
                
                if cleaned_data:
                    result.append({
                        "sheet_name": sheet_name,
                        "data": cleaned_data,
                        "row_count": len(cleaned_data),
                        "column_count": len(cleaned_data[0]) if cleaned_data else 0
                    })
            
            wb.close()
            
            logger.info("Excel file processed successfully", {
                "sheets": len(result),
                "total_rows": sum(sheet["row_count"] for sheet in result)
            })
            
            return result
            
        except Exception as error:
            logger.error("Error reading Excel file", {"error": str(error)})
            raise
    
    async def convert_to_chunks(
        self,
        excel_data: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Convert Excel data to text chunks for RAG.
        
        Args:
            excel_data: Extracted Excel data
            
        Returns:
            List of text chunks
        """
        try:
            chunks = []
            
            for sheet in excel_data:
                sheet_text = self._sheet_to_text(sheet)
                sheet_chunks = self._text_to_chunks(sheet_text, sheet["sheet_name"])
                chunks.extend(sheet_chunks)
            
            logger.info(f"Created {len(chunks)} text chunks from Excel data")
            return chunks
            
        except Exception as error:
            logger.error("Error converting Excel to chunks", {"error": str(error)})
            raise
    
    def _sheet_to_text(self, sheet: Dict[str, Any]) -> str:
        """
        Convert sheet data to formatted text.
        
        Args:
            sheet: Sheet data
            
        Returns:
            Formatted text
        """
        text = f"Sheet: {sheet['sheet_name']}\n\n"
        
        for row_idx, row in enumerate(sheet["data"]):
            if row_idx == 0:
                # Header row
                headers = [str(cell) if cell is not None else "" for cell in row]
                text += f"Headers: {' | '.join(headers)}\n"
                text += "=" * 80 + "\n"
            else:
                # Data rows
                row_data = []
                headers = sheet["data"][0]
                
                for cell_idx, cell in enumerate(row):
                    header = headers[cell_idx] if cell_idx < len(headers) else f"Column{cell_idx}"
                    cell_value = str(cell) if cell is not None else ""
                    row_data.append(f"{header}: {cell_value}")
                
                text += f"Row {row_idx}: {', '.join(row_data)}\n"
        
        return text
    
    def _text_to_chunks(
        self,
        text: str,
        source: str
    ) -> List[Dict[str, Any]]:
        """
        Split text into chunks with overlap.
        
        Args:
            text: Text to split
            source: Source identifier
            
        Returns:
            List of chunks
        """
        chunks = []
        paragraphs = text.split("\n\n")
        
        current_chunk = ""
        chunk_index = 0
        
        for paragraph in paragraphs:
            if len(current_chunk + paragraph) > self.chunk_size:
                if current_chunk:
                    chunks.append({
                        "id": f"{source}-chunk-{chunk_index}",
                        "source": source,
                        "content": current_chunk.strip(),
                        "metadata": {
                            "type": "excel",
                            "chunk_index": chunk_index,
                            "length": len(current_chunk)
                        }
                    })
                    chunk_index += 1
                    
                    # Create overlap
                    words = current_chunk.split()
                    overlap_words = words[-(self.chunk_overlap // 10):]
                    current_chunk = " ".join(overlap_words) + "\n\n" + paragraph
                else:
                    current_chunk = paragraph
            else:
                current_chunk += ("\n\n" if current_chunk else "") + paragraph
        
        # Add final chunk
        if current_chunk:
            chunks.append({
                "id": f"{source}-chunk-{chunk_index}",
                "source": source,
                "content": current_chunk.strip(),
                "metadata": {
                    "type": "excel",
                    "chunk_index": chunk_index,
                    "length": len(current_chunk)
                }
            })
        
        return chunks
    
    def _clean_data(self, data: List[tuple]) -> List[tuple]:
        """
        Clean and validate data.
        
        Args:
            data: Raw data from Excel
            
        Returns:
            Cleaned data
        """
        return [row for row in data if row and len(row) > 0]
    
    async def process(self, file_path: str) -> Dict[str, Any]:
        """
        Process Excel file and return chunks ready for RAG.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Processed data with chunks
        """
        excel_data = await self.read_excel(file_path)
        chunks = await self.convert_to_chunks(excel_data)
        
        return {
            "source": Path(file_path).name,
            "type": "excel",
            "chunks": chunks,
            "metadata": {
                "sheets": len(excel_data),
                "total_chunks": len(chunks)
            }
        }
